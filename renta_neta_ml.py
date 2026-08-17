# -*- coding: utf-8 -*-
"""
Renta Neta MercadoLibre + Canal Físico — ELDOM EL BAZAR / DELERAL S.A.
======================================================================
Herramienta de cierre mensual: subís los reportes de cada costo y la app
netea todo para darte la renta neta de MercadoLibre, la renta OPERATIVA
del canal físico (tienda) y la suma consolidada de ambas operativas.

Canales/fuentes que procesa:
  MERCADOLIBRE
  1. Reporte de Facturación de MercadoLibre (.xlsx)  -> comisión, envíos ME2/Colecta, publicidad, etc.
  2. Reporte de Notas de Crédito de MercadoLibre (.xlsx)
  3. Reporte de Costeo FacturApp - FACTURAS (.pdf, medio de pago M LIBRE)
  4. Reporte de Costeo FacturApp - NOTAS DE CRÉDITO (.pdf, medio de pago M LIBRE)
  5. Facturación mensual DAC (.pdf)                  -> flete ME1 (cuenta corriente)
  6. Nota de Crédito DAC / bonificación (.pdf)       -> descuento 20%
  7. Factura FLEX Distrilogic (.pdf)                 -> logística FLEX

  CANAL FÍSICO (tienda, todo lo que NO es M LIBRE)
  8. Reporte de Costeo FacturApp - TODOS los métodos (.pdf)  -> costo y venta por documento
  9. Reporte de Ventas - Medios de pago (.pdf)               -> forma de pago por venta
 10. Reporte de Notas de Crédito - Medios de pago (.pdf)     -> devoluciones (opcional)

El canal físico calcula la renta OPERATIVA por forma de pago:
    operativa = venta − costo − cobranza      (cobranza = venta × tasa)
Tasas de cobranza: Débito 1,35% · Efectivo 0% · Transferencia (TR BROU) 0% ·
resto de las tarjetas 8,36%. (Editable en MAPEO_FIS / TASAS_FIS.)
Todo en peso-equivalente (usa la columna Venta $ del costeo, ya convertida al TC diario).

Deploy: Streamlit Community Cloud o Render. Solo requiere renta_neta_ml.py + requirements.txt.
"""

import io
import re
import pandas as pd
import pdfplumber
import streamlit as st
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

IVA_RATE = 0.22

# ===========================================================================
#  CANAL FÍSICO — tasas de cobranza y mapeo de formas de pago
#  EDITAR ACÁ si alguna forma de pago va a otra tasa.
#  Regla vigente: todo 8,36% (crédito) salvo Débito (1,35%), EFECTIVO y TR BROU (0%).
#  M LIBRE se excluye siempre (es el canal ML, no físico).
# ===========================================================================
TASAS_FIS = {"debito": 0.0135, "credito": 0.0836, "efectivo": 0.0, "transferencia": 0.0}

MAPEO_FIS = {
    "Debito":             "debito",         # 1,35%
    "EFECTIVO":           "efectivo",       # 0%
    "TR BROU":            "transferencia",  # 0%
    # el resto cae por defecto en "credito" (8,36%):
    "VISA":               "credito",
    "MASTER CARD":        "credito",
    "WEB MasterCard":     "credito",
    "OCA":                "credito",
    "CABAL":              "credito",
    "CREDITEL":           "credito",
    "CREDITOS DIRECTOS":  "credito",
    "PASS CARD":          "credito",
    "M PAGO":             "credito",
    "CREDITO DE LA CASA": "credito",
}
TASA_DEFECTO_FIS = "credito"   # cualquier forma de pago no listada -> 8,36%

METODOS_FIS = ["M LIBRE", "M PAGO", "WEB MasterCard", "MASTER CARD", "CREDITOS DIRECTOS",
               "CREDITO DE LA CASA", "CREDITEL", "PASS CARD", "TR BROU", "Debito",
               "EFECTIVO", "VISA", "CABAL", "OCA"]


# ---------------------------------------------------------------------------
# Utilidades de parseo numérico
# ---------------------------------------------------------------------------
def parse_amount(s):
    """Convierte un texto de importe a float, tolerando formato UY (1.234,56)
    y americano (1234.56)."""
    s = str(s).replace("$", "").replace("UYU", "").replace("\u00a0", " ").strip()
    s = s.replace(" ", "")
    if not s:
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):          # coma decimal (UY)
            s = s.replace(".", "").replace(",", ".")
        else:                                     # punto decimal, coma miles
            s = s.replace(",", "")
    elif "," in s:                                # solo coma -> decimal
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _pdf_text(f):
    f.seek(0)
    out = []
    with pdfplumber.open(f) as pdf:
        for pg in pdf.pages:
            out.append(pg.extract_text() or "")
    return "\n".join(out)


# ---------------------------------------------------------------------------
# 1-2) Cargos de MercadoLibre (facturación + notas de crédito, .xlsx)
# ---------------------------------------------------------------------------
def parse_ml_charges(fac_file, nc_file):
    frames = []
    for f in (fac_file, nc_file):
        if f is None:
            continue
        f.seek(0)
        df = pd.read_excel(f, sheet_name="REPORT", header=7).dropna(how="all")
        frames.append(df)
    if not frames:
        return None
    allml = pd.concat(frames, ignore_index=True)
    d = allml["Detalle"].astype(str)
    val = pd.to_numeric(allml["Valor del cargo"], errors="coerce").fillna(0)

    def s(mask):
        return round(float(val[mask].sum()), 2)

    return dict(
        comision=s(d.str.contains("cargo por venta", case=False, na=False)),
        envios=s(d.str.contains("envíos de Mercado Libre", case=False, na=False)
                 | d.str.contains("costo de envío", case=False, na=False)),
        prodads=s(d.str.contains("Product Ads", case=False, na=False)),
        dispads=s(d.str.contains("Display Ads", case=False, na=False)),
        aseso=s(d.str.contains("Asesoría", case=False, na=False)),
        manten=s(d.str.contains("mantenimiento de Mi página", case=False, na=False)),
        devfee=s(d.str.contains("Cargo por devolución", case=False, na=False)),
    )


# ---------------------------------------------------------------------------
# 3-4) Costeo FacturApp (.pdf) -> (costo, venta) con signo del reporte  [BLOQUE ML]
# ---------------------------------------------------------------------------
def parse_costeo_pdf(f):
    if f is None:
        return (0.0, 0.0)
    text = _pdf_text(f)
    # Renglón oficial de totales: "TOTAL: costo$ costoU$S venta$ ventaU$S"
    summary = None
    for ln in text.split("\n"):
        m = re.search(
            r"TOTAL:\s*(-?\d+\.\d{2})\s+(-?\d+\.\d{2})\s+(-?\d+\.\d{2})\s+(-?\d+\.\d{2})",
            ln.strip())
        if m:
            summary = m
    if summary:
        return (float(summary.group(1)), float(summary.group(3)))
    # Fallback: sumar líneas de detalle
    costo = venta = 0.0
    pat = re.compile(
        r"(FACTURA CONTADO|DEVOLUCION CONTAD)\s+\S+.*?\$?\s*[\d.]+\s+\$?\s*"
        r"(-?\d+\.\d+)\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)")
    for ln in text.split("\n"):
        m = pat.search(ln)
        if m:
            costo += float(m.group(2))
            venta += float(m.group(4))
    return (round(costo, 2), round(venta, 2))


# ---------------------------------------------------------------------------
# 5) Facturación DAC (.pdf) -> total, salida (ventas ME1), entrada (proveedores)
# ---------------------------------------------------------------------------
def parse_dac_fact(f):
    if f is None:
        return dict(total=0.0, salida=0.0, entrada=0.0)
    total = sal = ent = 0.0
    f.seek(0)
    with pdfplumber.open(f) as pdf:
        for pg in pdf.pages:
            for ln in (pg.extract_text() or "").split("\n"):
                s = ln.strip()
                is_sal = "Cta.Cte. Remitente" in s
                is_ent = "Flete destino" in s
                if not (is_sal or is_ent):
                    continue
                nums = re.findall(r"\d[\d.]*,\d\d", s)
                if not nums:
                    continue
                v = parse_amount(nums[-1])
                total += v
                if is_sal:
                    sal += v
                else:
                    ent += v
    return dict(total=round(total, 2), salida=round(sal, 2), entrada=round(ent, 2))


# ---------------------------------------------------------------------------
# 6) Nota de crédito DAC (bonificación) -> total (con IVA), subtotal (sin IVA)
# ---------------------------------------------------------------------------
def parse_dac_nc(f):
    if f is None:
        return dict(total=0.0, subtotal=0.0)
    text = _pdf_text(f)
    total = sub = 0.0
    m = re.search(r"TOTAL:\s*\$?\s*([\d.,]+)", text)
    if m:
        total = parse_amount(m.group(1))
    m = re.search(r"Subtotal:\s*\$?\s*([\d.,]+)", text)
    if m:
        sub = parse_amount(m.group(1))
    return dict(total=round(total, 2), subtotal=round(sub, 2))


# ---------------------------------------------------------------------------
# 7) Factura FLEX Distrilogic -> total (con IVA), neto (sin IVA)
# ---------------------------------------------------------------------------
def parse_flex(f):
    if f is None:
        return dict(total=0.0, neto=0.0)
    text = _pdf_text(f)
    total = neto = 0.0
    m = re.search(r"TOTAL A PAGAR\s*\n?\s*(?:UYU)?\s*([\d.,]+)", text)
    if not m:
        m = re.search(r"TOTAL A PAGAR[^\d]*([\d.,]+)", text)
    if m:
        total = parse_amount(m.group(1))
    m2 = re.search(r"([\d.]+,\d\d)\s+([\d.]+,\d\d)\s+([\d.]+,\d\d)", text)
    if m2:
        neto = parse_amount(m2.group(1))
    if not neto and total:
        neto = round(total / (1 + IVA_RATE), 2)
    return dict(total=round(total, 2), neto=round(neto, 2))


# ===========================================================================
#  8-10) CANAL FÍSICO — parsers y cálculo
# ===========================================================================
def parse_costeo_detalle(f):
    """Costeo (todos los métodos) -> {documento: (costo$, venta$)} peso-equiv."""
    if f is None:
        return {}
    text = _pdf_text(f)
    pat = re.compile(r"FACTURA CONTADO (\d+) \$ [\d.]+ (?:U\$S|\$) ([\d.]+) [\d.]+ ([\d.]+) [\d.]+")
    return {m.group(1): (float(m.group(2)), float(m.group(3))) for m in pat.finditer(text)}


def parse_ventas_formapago(f):
    """Ventas medios de pago -> {documento: [(metodo, monto, moneda), ...]}
    (soporta pagos partidos en varias formas de pago)."""
    if f is None:
        return {}
    text = _pdf_text(f)
    mpat = re.compile(r"^(" + "|".join(re.escape(x) for x in METODOS_FIS) +
                      r")\s+([\d.]+)\s+(U\$S|\$)\s*$", re.M)
    hdr = re.compile(r"FACTURA CONTADO (\d+) e(?:Ticket|Factura)")
    starts = [(m.start(), m.group(1)) for m in hdr.finditer(text)]
    res = {}
    for i, (s, doc) in enumerate(starts):
        e = starts[i + 1][0] if i + 1 < len(starts) else len(text)
        res.setdefault(doc, [])
        for mm in mpat.finditer(text[s:e]):
            res[doc].append((mm.group(1), float(mm.group(2)), mm.group(3)))
    return res


def parse_nc_formapago(f):
    """NC medios de pago -> [(metodo, monto)] de las devoluciones NO M LIBRE (físicas)."""
    if f is None:
        return []
    text = _pdf_text(f)
    mpat = re.compile(r"^(" + "|".join(re.escape(x) for x in METODOS_FIS) +
                      r")\s+([\d.]+)\s+(?:U\$S|\$)\s*$", re.M)
    return [(mm.group(1), float(mm.group(2))) for mm in mpat.finditer(text)
            if mm.group(1) != "M LIBRE"]


def _tasa_fis(metodo):
    return TASAS_FIS[MAPEO_FIS.get(metodo, TASA_DEFECTO_FIS)]


def compute_fisico(costeo_det, ventas_fp, nc_fis):
    """Renta operativa del canal físico por forma de pago (excluye M LIBRE).
    Las NC físicas (devoluciones totales) se netean reversando venta y costo
    del documento original, matcheado por (forma de pago, monto)."""
    venta = defaultdict(float)
    costo = defaultdict(float)
    n = defaultdict(int)
    fact_fis = fact_ml = 0.0   # facturación bruta (con IVA, antes de NC)
    for doc, lst in ventas_fp.items():
        if doc not in costeo_det:
            continue
        c_tot, v_tot = costeo_det[doc]
        # facturación bruta por canal (cada documento es puro M LIBRE o puro físico)
        if any(m == "M LIBRE" for m, _, _ in lst):
            fact_ml += v_tot
        else:
            fact_fis += v_tot
        tot = sum(a for _, a, _ in lst)
        for meth, amt, _cur in lst:
            if meth == "M LIBRE":
                continue
            p = (amt / tot) if tot > 0 else (1.0 / len(lst))
            venta[meth] += v_tot * p
            costo[meth] += c_tot * p
            n[meth] += 1

    # NC físicas: reverso venta+costo del documento original (devolución total)
    idx = defaultdict(list)
    for doc, lst in ventas_fp.items():
        for meth, amt, _cur in lst:
            if meth != "M LIBRE":
                idx[(meth, round(amt, 1))].append(doc)
    usados = set()
    nc_rev = []
    for meth, monto in nc_fis:
        cand = [d for d in idx.get((meth, round(monto, 1)), []) if d not in usados]
        if cand and cand[0] in costeo_det:
            d = cand[0]
            usados.add(d)
            c_tot, v_tot = costeo_det[d]
            venta[meth] -= v_tot
            costo[meth] -= c_tot
            nc_rev.append((meth, d, v_tot))
        else:
            nc_rev.append((meth, None, monto))   # no matcheada (revisar manual)

    filas = []
    tv = tc = tcob = 0.0
    for meth in sorted(venta, key=lambda x: -venta[x]):
        v = venta[meth]
        c = costo[meth]
        tasa = _tasa_fis(meth)
        cob = v * tasa
        filas.append(dict(metodo=meth, n=n[meth], venta=v, costo=c,
                          tasa=tasa, cobranza=cob, operativa=v - c - cob))
        tv += v
        tc += c
        tcob += cob

    return dict(
        filas=filas, total_venta=round(tv, 2), total_costo=round(tc, 2),
        total_cobranza=round(tcob, 2), margen_bruto=round(tv - tc, 2),
        operativa=round(tv - tc - tcob, 2), nc_rev=nc_rev,
        facturacion=round(fact_fis, 2), facturacion_ml=round(fact_ml, 2),
        facturacion_total=round(fact_ml + fact_fis, 2),
    )


# ---------------------------------------------------------------------------
# Cálculo del estado de resultados (BLOQUE ML)
# ---------------------------------------------------------------------------
def compute(ml, cost_fac, cost_nc, dac, dacnc, flex,
            impuestos=0.0, base_sin_iva=True, coef_irae=0.0181, iva_a_pagar=0.0):
    venta_bruta, venta_dev = cost_fac[1], cost_nc[1]
    costo_bruto, costo_dev = cost_fac[0], cost_nc[0]
    venta_neta = round(venta_bruta + venta_dev, 2)
    costo_neto = round(costo_bruto + costo_dev, 2)
    gb = round(venta_neta - costo_neto, 2)

    ml = ml or dict(comision=0, envios=0, prodads=0, dispads=0, aseso=0, manten=0, devfee=0)
    cargos_ml = round(ml["comision"] + ml["prodads"] + ml["dispads"]
                      + ml["aseso"] + ml["manten"] + ml["devfee"], 2)

    dac_neto = round(dac["total"] - dacnc["total"], 2)          # con IVA
    envios = round(ml["envios"] + dac_neto + flex["total"], 2)

    rno = round(gb - cargos_ml - envios, 2)                     # operativa con IVA
    margen = (rno / venta_neta) if venta_neta else 0.0

    # Cierre económico: base sin IVA (÷1,22) + IRAE por ficto
    venta_siva = round(venta_neta / (1 + IVA_RATE), 2)
    if base_sin_iva:
        rno_base = round(rno / (1 + IVA_RATE), 2)
    else:
        rno_base = rno
    irae = round(venta_siva * coef_irae, 2)
    final = round(rno_base - irae - iva_a_pagar - impuestos, 2)

    return dict(
        venta_bruta=venta_bruta, venta_dev=venta_dev, venta_neta=venta_neta,
        costo_bruto=costo_bruto, costo_dev=costo_dev, costo_neto=costo_neto,
        gb=gb, cargos_ml=cargos_ml, dac_neto=dac_neto, flex=flex["total"],
        me2=ml["envios"], envios=envios, rno=rno, margen=margen, ml=ml,
        dac=dac, dacnc=dacnc,
        venta_siva=venta_siva, rno_base=rno_base, irae=irae,
        iva_a_pagar=iva_a_pagar, final=final,
    )


# ---------------------------------------------------------------------------
# Exportación a Excel con formato
# ---------------------------------------------------------------------------
def build_excel(r, fis=None, op_total=None):
    MONEY = "#,##0.00;(#,##0.00)"; PCT = "0.0%"
    BLUE = Font(name="Arial", color="0000FF", size=10)
    BLACK = Font(name="Arial", color="000000", size=10)
    BOLD = Font(name="Arial", bold=True, size=10)
    BOLDW = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    TITLE = Font(name="Arial", bold=True, size=13, color="7B1E22")
    CRIMSON = PatternFill("solid", fgColor="7B1E22")
    GREEN = PatternFill("solid", fgColor="E2EFDA")
    GOLD = PatternFill("solid", fgColor="FFF2CC")
    top = Border(top=Side(style="thin", color="000000"))
    dbl = Border(top=Side(style="thin", color="000000"),
                 bottom=Side(style="double", color="000000"))

    wb = Workbook(); ws = wb.active; ws.title = "MercadoLibre"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 20
    ws["B2"] = "ELDOM EL BAZAR — DELERAL S.A."; ws["B2"].font = TITLE
    ws["B3"] = "Renta neta operación MercadoLibre"; ws["B3"].font = BOLD

    row = [5]
    def line(label, value, font=BLACK, fmt=MONEY, fill=None, bd=None, sheet=None):
        w = sheet or ws
        rr = row[0]
        c1 = w.cell(rr, 2, label); c1.font = font
        c2 = w.cell(rr, 3, value); c2.font = font; c2.number_format = fmt
        c2.alignment = Alignment(horizontal="right")
        if fill:
            c1.fill = fill; c2.fill = fill
        if bd:
            c2.border = bd
        row[0] += 1
    def blank():
        row[0] += 1
    def hdr(t, sheet=None):
        w = sheet or ws
        rr = row[0]
        w.cell(rr, 2, t).font = BOLDW
        w.cell(rr, 2).fill = CRIMSON; w.cell(rr, 3).fill = CRIMSON
        row[0] += 1

    hdr("INGRESOS")
    line("Ventas facturadas (FacturApp)", r["venta_bruta"], BLUE)
    line("(–) Devoluciones / Notas de crédito", r["venta_dev"], BLUE)
    line("Ventas netas", r["venta_neta"], BOLD, bd=top); blank()
    hdr("COSTO DE MERCADERÍA")
    line("Costo de productos vendidos", -abs(r["costo_bruto"]), BLUE)
    line("(+) Reversión por devoluciones", abs(r["costo_dev"]), BLUE)
    line("Costo neto", -abs(r["costo_neto"]), BOLD, bd=top); blank()
    line("GANANCIA BRUTA", r["gb"], BOLD, MONEY, GREEN)
    line("Margen bruto", r["gb"] / r["venta_neta"] if r["venta_neta"] else 0, BLACK, PCT); blank()
    hdr("CARGOS DE MERCADOLIBRE")
    line("Comisión por venta (neta)", -r["ml"]["comision"], BLUE)
    line("Publicidad — Product Ads", -r["ml"]["prodads"], BLUE)
    line("Publicidad — Display Ads", -r["ml"]["dispads"], BLUE)
    line("Asesoría Comercial", -r["ml"]["aseso"], BLUE)
    line("Mantenimiento Mi página (neto)", -r["ml"]["manten"], BLUE)
    line("Cargo por devolución (fee ML)", -r["ml"]["devfee"], BLUE)
    line("Subtotal cargos ML", -r["cargos_ml"], BOLD, bd=top); blank()
    hdr("ENVÍOS (3 canales)")
    line("ME2 / Colecta (MercadoLibre)", -r["me2"], BLUE)
    line("ME1 / DAC (neto cta.cte., con IVA)", -r["dac_neto"], BLUE)
    line("FLEX / Distrilogic (con IVA)", -r["flex"], BLUE)
    line("Subtotal envíos", -r["envios"], BOLD, bd=top); blank()
    line("RENTA NETA OPERATIVA (con IVA)", r["rno"], BOLDW, MONEY, CRIMSON)
    line("Margen operativo", r["margen"], BLACK, PCT); blank()
    hdr("CIERRE ECONÓMICO (base sin IVA + IRAE)")
    line("Renta operativa base sin IVA", r["rno_base"], BOLD)
    line("(–) IRAE atribuible ML", -r["irae"], BLUE)
    line("(–) IVA neto a pagar DGI", -r["iva_a_pagar"], BLUE)
    line("RENTA NETA FINAL", r["final"], BOLDW, MONEY, CRIMSON, bd=dbl)

    # ---------------- Hoja Canal físico + consolidado ----------------
    if fis is not None:
        ws2 = wb.create_sheet("Canal físico")
        ws2.sheet_view.showGridLines = False
        for col, w in (("A", 2), ("B", 26), ("C", 16), ("D", 16),
                       ("E", 10), ("F", 15), ("G", 16)):
            ws2.column_dimensions[col].width = w
        ws2["B2"] = "ELDOM EL BAZAR — DELERAL S.A."; ws2["B2"].font = TITLE
        ws2["B3"] = "Canal físico (tienda) — renta operativa por forma de pago"; ws2["B3"].font = BOLD
        headers = ["Forma de pago", "Venta $", "Costo $", "Tasa", "Cobranza $", "Operativa $"]
        hr = 5
        for j, h in enumerate(headers):
            c = ws2.cell(hr, 2 + j, h); c.font = BOLDW; c.fill = CRIMSON
            c.alignment = Alignment(horizontal="right" if j else "left")
        rr = hr + 1
        for f in fis["filas"]:
            ws2.cell(rr, 2, f["metodo"]).font = BLACK
            for j, (val, fmt) in enumerate([
                (f["venta"], MONEY), (f["costo"], MONEY), (f["tasa"], PCT),
                (f["cobranza"], MONEY), (f["operativa"], MONEY)]):
                c = ws2.cell(rr, 3 + j, val); c.font = BLACK; c.number_format = fmt
                c.alignment = Alignment(horizontal="right")
            rr += 1
        # totales
        for j, (val, fmt) in enumerate([
            (fis["total_venta"], MONEY), (fis["total_costo"], MONEY), (None, PCT),
            (fis["total_cobranza"], MONEY), (fis["operativa"], MONEY)]):
            lab = ws2.cell(rr, 2, "TOTAL FÍSICO"); lab.font = BOLD
            if val is not None:
                c = ws2.cell(rr, 3 + j, val); c.font = BOLD; c.number_format = fmt
                c.border = top; c.alignment = Alignment(horizontal="right")

        # ---------------- Hoja Consolidado (facturación + operativa + % por canal) ----------------
        NOTE = Font(name="Arial", size=8, italic=True, color="808080")
        op_ml = r["rno"]; op_fis = fis["operativa"]
        op_tot = op_total if op_total is not None else round(op_ml + op_fis, 2)
        f_ml = fis.get("facturacion_ml", r["venta_bruta"]) or 0.0
        f_fis = fis.get("facturacion", 0.0) or 0.0
        f_tot = fis.get("facturacion_total", round(f_ml + f_fis, 2)) or 0.0
        pct = lambda o, f: (o / f) if f else 0.0

        ws3 = wb.create_sheet("Consolidado")
        ws3.sheet_view.showGridLines = False
        for col, w in (("A", 2), ("B", 30), ("C", 18), ("D", 18), ("E", 12)):
            ws3.column_dimensions[col].width = w
        ws3["B2"] = "ELDOM EL BAZAR — DELERAL S.A."; ws3["B2"].font = TITLE
        ws3["B3"] = "Renta operativa consolidada por canal (con IVA)"; ws3["B3"].font = BOLD
        for j, h in enumerate(["Canal", "Facturación $", "Renta operativa $", "% Renta"]):
            c = ws3.cell(5, 2 + j, h); c.font = BOLDW; c.fill = CRIMSON
            c.alignment = Alignment(horizontal="right" if j else "left")
        data = [
            ("MercadoLibre", f_ml, op_ml, pct(op_ml, f_ml), BLACK, None),
            ("Canal físico", f_fis, op_fis, pct(op_fis, f_fis), BLACK, None),
            ("TOTAL", f_tot, op_tot, pct(op_tot, f_tot), BOLD, GOLD),
        ]
        rr = 6
        for canal, fac, ope, pc, fnt, fill in data:
            cc = ws3.cell(rr, 2, canal); cc.font = fnt
            if fill:
                cc.fill = fill
            for j, (val, fmt) in enumerate([(fac, MONEY), (ope, MONEY), (pc, "0.00%")]):
                c = ws3.cell(rr, 3 + j, val); c.font = fnt; c.number_format = fmt
                c.alignment = Alignment(horizontal="right")
                if fill:
                    c.fill = fill
                if canal == "TOTAL":
                    c.border = dbl
            rr += 1
        for i, t in enumerate([
            "Facturación = ventas brutas facturadas con IVA (reporte de costeo, antes de NC). Los tres canales suman el total del reporte.",
            "Renta operativa ML: neta de comisión, publicidad (ADS), asesoría, mantenimiento, devoluciones y envíos ME2/ME1/FLEX.",
            "Renta operativa física: neta de cobranza de tarjeta. % Renta = Renta operativa / Facturación de cada canal."]):
            ws3.cell(rr + 1 + i, 2, t).font = NOTE

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio.getvalue()


# ---------------------------------------------------------------------------
# Interfaz
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Renta Neta ML — ELDOM", page_icon="📊", layout="wide")
st.title("📊 Renta Neta MercadoLibre + Canal Físico")
st.caption("ELDOM EL BAZAR / DELERAL S.A. — cierre mensual. Subí cada reporte y presioná Procesar.")

# ---- Estilos de marca (colores casilleros / botón / títulos) ----
st.markdown("""
<style>
h2, h3 { color:#7B1E22 !important; }
[data-testid="stFileUploaderDropzone"]{
  background:#FBF6F2; border:1.6px dashed #7B1E22; border-radius:12px;
}
[data-testid="stFileUploaderDropzone"]:hover{ background:#F3E4DD; border-color:#651519; }
div.stButton > button[kind="primary"]{
  background:#7B1E22; color:#fff; border:0; font-weight:700; border-radius:10px;
}
div.stButton > button[kind="primary"]:hover{ background:#651519; color:#fff; }
</style>
""", unsafe_allow_html=True)

with st.expander("¿Qué archivo va en cada casillero?"):
    st.markdown("""
**MercadoLibre**
- **Facturación ML** y **Notas de Crédito ML** (`.xlsx`): reportes de facturación de MercadoLibre. De ahí salen comisión, envíos ME2/Colecta, publicidad, asesoría, etc.
- **Costeo FacturApp – Facturas** y **– Notas de Crédito** (`.pdf`): costeo de FacturApp **filtrado a medio de pago M LIBRE**. De ahí salen la venta y el costo del producto de ML.
- **DAC Facturación** y **DAC Nota de Crédito** (`.pdf`): flete ME1 por cuenta corriente y su bonificación del 20%.
- **FLEX Distrilogic** (`.pdf`): factura de logística FLEX.

**Canal físico (tienda)**
- **Costeo – TODOS los métodos** (`.pdf`): costeo de FacturApp **sin filtrar por medio de pago**. Trae costo y venta por documento.
- **Ventas – Medios de pago** (`.pdf`): listado de ventas con la forma de pago de cada una. Se excluye M LIBRE y se desglosa el resto.
- **NC – Medios de pago** (`.pdf`, opcional): devoluciones. Las NC físicas se netean automáticamente reversando venta y costo del documento original.

Podés subir solo los que tengas; lo que falte queda en cero. El canal físico necesita al menos el costeo de todos los métodos + el de ventas medios de pago.
""")

c1, c2 = st.columns(2)
with c1:
    st.subheader("MercadoLibre")
    ml_fac = st.file_uploader(
        "Facturación ML (.xlsx)", type=["xlsx"], key="mlf",
        help="Buscá el archivo:  Reporte_Facturacion_MercadoLibre_<Mes><Año>.xlsx"
             "  ·  ej: Reporte_Facturacion_MercadoLibre_Jun2026.xlsx")
    ml_nc = st.file_uploader(
        "Notas de Crédito ML (.xlsx)", type=["xlsx"], key="mln",
        help="Buscá el archivo:  Reporte_Notas_Credito_MercadoLibre_<Mes><Año>.xlsx"
             "  ·  ej: Reporte_Notas_Credito_MercadoLibre_Jun2026.xlsx")
    st.subheader("Costeo FacturApp (M LIBRE)")
    co_fac = st.file_uploader(
        "Costeo – Facturas (.pdf)", type=["pdf"], key="cof",
        help="Reporte de costeo de FacturApp FILTRADO a medio de pago M LIBRE — Facturas (.pdf)")
    co_nc = st.file_uploader(
        "Costeo – Notas de Crédito (.pdf)", type=["pdf"], key="con",
        help="Reporte de costeo de FacturApp FILTRADO a M LIBRE — Notas de Crédito / Devoluciones (.pdf)")
with c2:
    st.subheader("Flete DAC (ME1)")
    dac_f = st.file_uploader(
        "DAC – Facturación mensual (.pdf)", type=["pdf"], key="dacf",
        help="Buscá el archivo:  Facturacion_Mensual-INV_<número>.pdf  (reporte de facturación mensual de DAC)")
    dac_n = st.file_uploader(
        "DAC – Nota de Crédito / bonificación (.pdf)", type=["pdf"], key="dacn",
        help="Buscá el archivo:  Factura-NC_<número>.pdf  (nota de crédito / bonificación mensual de DAC)")
    st.subheader("Logística FLEX")
    flex_f = st.file_uploader(
        "Distrilogic – Factura (.pdf)", type=["pdf"], key="flexf",
        help="Factura de Distrilogic FLEX (.pdf). Si viene como reporte de servicios "
             "(ELDOM_SERVICIOS_DE_<Mes>_<Año>.xlsx), cargá el total FLEX de otra forma / avisá.")

st.divider()
st.subheader("🏬 Canal físico (tienda)")
cf1, cf2, cf3 = st.columns(3)
with cf1:
    fis_costeo = st.file_uploader(
        "Costeo – TODOS los métodos (.pdf)", type=["pdf"], key="fisc",
        help="Buscá el archivo:  total_<mes>_<año>_todos_metodos_de_pago_Reporte_de_costeo.pdf")
with cf2:
    fis_venta = st.file_uploader(
        "Ventas – Medios de pago (.pdf)", type=["pdf"], key="fisv",
        help="Buscá el archivo:  venta_total_<mes>_<año>.pdf  (listado de ventas con forma de pago)")
with cf3:
    fis_nc = st.file_uploader(
        "NC – Medios de pago (.pdf, opcional)", type=["pdf"], key="fisn",
        help="Buscá el archivo:  NC_total_<mes>_<año>.pdf  (devoluciones con forma de pago)")

st.divider()
st.subheader("Cierre de impuestos (solo aplica al bloque ML)")
ci1, ci2, ci3 = st.columns(3)
with ci1:
    base_sin_iva = st.checkbox(
        "Llevar a base sin IVA (÷1,22)", value=True,
        help="Costo y venta están cargados con IVA. El IVA es pass-through; se elimina dividiendo por 1,22.")
with ci2:
    coef_irae = st.number_input(
        "Coeficiente IRAE (ficto)", min_value=0.0, value=0.0181, step=0.0001, format="%.4f",
        help="Del reporte del contador. IRAE = ventas sin IVA × coeficiente.")
with ci3:
    iva_a_pagar = st.number_input(
        "IVA neto a pagar DGI", min_value=0.0, value=0.0, step=1000.0, format="%.2f",
        help="Según el contador. Si tenés crédito acumulado, es 0.")

if st.button("⚙️  Procesar", type="primary", use_container_width=True):
    try:
        with st.spinner("⚙️  Procesando reportes… un momento"):
            ml = parse_ml_charges(ml_fac, ml_nc)
            cost_fac = parse_costeo_pdf(co_fac)
            cost_nc = parse_costeo_pdf(co_nc)
            dac = parse_dac_fact(dac_f)
            dacnc = parse_dac_nc(dac_n)
            flex = parse_flex(flex_f)
            r = compute(ml, cost_fac, cost_nc, dac, dacnc, flex,
                        base_sin_iva=base_sin_iva, coef_irae=coef_irae, iva_a_pagar=iva_a_pagar)

            # ---- Canal físico ----
            costeo_det = parse_costeo_detalle(fis_costeo)
            ventas_fp = parse_ventas_formapago(fis_venta)
            nc_fis = parse_nc_formapago(fis_nc)
            fis = None
            op_total = None
            if costeo_det and ventas_fp:
                fis = compute_fisico(costeo_det, ventas_fp, nc_fis)
                op_total = round(r["rno"] + fis["operativa"], 2)

        warns = []
        if dacnc["total"] > 0 and dac["total"] == 0:
            warns.append("Cargaste la **Nota de Crédito de DAC** pero falta la **Facturación mensual de DAC**. "
                         "El flete ME1 quedó como ingreso en vez de costo — subí la facturación de DAC y volvé a procesar.")
        if r["dac_neto"] < 0:
            warns.append("El neto de DAC dio negativo. Revisá que los dos PDF de DAC estén en los casilleros correctos.")
        if flex["total"] > 0 and r["me2"] == 0 and dac["total"] == 0:
            warns.append("Solo se leyó FLEX en envíos. Verificá que cargaste también los reportes de ML y DAC.")
        # Guards del canal físico
        if fis_venta is not None and not costeo_det:
            warns.append("Subiste **Ventas – Medios de pago** pero no se leyó el **Costeo TODOS los métodos**. "
                         "El canal físico necesita ambos.")
        if fis is not None and fis["total_venta"] < 1000:
            warns.append("El canal físico dio casi cero. Puede que hayas subido el costeo **filtrado a M LIBRE** "
                         "en el casillero de TODOS los métodos — el físico son las ventas que NO son M LIBRE.")
        if fis is not None:
            no_match = [x for x in fis["nc_rev"] if x[1] is None]
            if no_match:
                warns.append("Algunas NC físicas no se pudieron matchear a su venta original "
                             f"({len(no_match)}). Puede ser una devolución parcial; revisá manualmente o subí el costeo de NC.")
        for w in warns:
            st.warning(w)

        # -------- Métricas --------
        if fis is not None:
            f_tot_m = fis.get("facturacion_total", 0.0) or 0.0
            pct_tot = (op_total / f_tot_m) if f_tot_m else 0.0
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Operativa ML (c/IVA)", f"$ {r['rno']:,.0f}", f"{r['margen']*100:.1f}% s/ventas")
            m2.metric("Operativa física (c/IVA)", f"$ {fis['operativa']:,.0f}")
            m3.metric("OPERATIVA TOTAL", f"$ {op_total:,.0f}", f"{pct_tot*100:.2f}% s/facturación")
            m4.metric("Renta neta FINAL ML", f"$ {r['final']:,.0f}")
        else:
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Ganancia bruta (c/IVA)", f"$ {r['gb']:,.0f}")
            m2.metric("Renta operativa (c/IVA)", f"$ {r['rno']:,.0f}", f"{r['margen']*100:.1f}% s/ventas")
            m3.metric("IRAE atribuible ML", f"$ {r['irae']:,.0f}")
            m4.metric("Renta neta FINAL", f"$ {r['final']:,.0f}")

        # -------- Estado de resultados ML --------
        st.markdown("#### MercadoLibre — estado de resultados")
        tabla = pd.DataFrame([
            ("Ventas netas (con IVA)", r["venta_neta"]),
            ("Costo neto de mercadería", -abs(r["costo_neto"])),
            ("= Ganancia bruta (con IVA)", r["gb"]),
            ("Comisión ML (neta)", -r["ml"]["comision"]),
            ("Publicidad Product Ads", -r["ml"]["prodads"]),
            ("Publicidad Display Ads", -r["ml"]["dispads"]),
            ("Asesoría Comercial", -r["ml"]["aseso"]),
            ("Mantenimiento Mi página", -r["ml"]["manten"]),
            ("Cargo por devolución", -r["ml"]["devfee"]),
            ("Envío ME2 / Colecta (ML)", -r["me2"]),
            ("Envío ME1 / DAC (neto)", -r["dac_neto"]),
            ("Envío FLEX / Distrilogic", -r["flex"]),
            ("= Renta operativa (con IVA)", r["rno"]),
            ("Renta operativa base sin IVA (÷1,22)", r["rno_base"]),
            ("(–) IRAE atribuible ML", -r["irae"]),
            ("(–) IVA neto a pagar DGI", -r["iva_a_pagar"]),
            ("= RENTA NETA FINAL", r["final"]),
        ], columns=["Concepto", "UYU"])
        st.dataframe(
            tabla.style.format({"UYU": "{:,.2f}"}),
            use_container_width=True, hide_index=True, height=560)

        # -------- Canal físico --------
        if fis is not None:
            st.markdown("#### Canal físico — renta operativa por forma de pago")
            tfis = pd.DataFrame([
                {"Forma de pago": f["metodo"], "Venta $": f["venta"], "Costo $": f["costo"],
                 "Tasa": f["tasa"], "Cobranza $": f["cobranza"], "Operativa $": f["operativa"]}
                for f in fis["filas"]
            ] + [{"Forma de pago": "TOTAL FÍSICO", "Venta $": fis["total_venta"],
                  "Costo $": fis["total_costo"], "Tasa": None,
                  "Cobranza $": fis["total_cobranza"], "Operativa $": fis["operativa"]}])
            st.dataframe(
                tfis.style.format({"Venta $": "{:,.2f}", "Costo $": "{:,.2f}",
                                   "Tasa": "{:.2%}", "Cobranza $": "{:,.2f}",
                                   "Operativa $": "{:,.2f}"}, na_rep=""),
                use_container_width=True, hide_index=True)

            st.markdown("#### Consolidado — facturación, renta operativa y % por canal")
            f_ml = fis.get("facturacion_ml", r["venta_bruta"]) or 0.0
            f_fis = fis.get("facturacion", 0.0) or 0.0
            f_tot = fis.get("facturacion_total", round(f_ml + f_fis, 2)) or 0.0
            pct = lambda o, f: (o / f) if f else 0.0
            tcons = pd.DataFrame([
                {"Canal": "MercadoLibre", "Facturación $": f_ml,
                 "Renta operativa $": r["rno"], "% Renta": pct(r["rno"], f_ml)},
                {"Canal": "Canal físico", "Facturación $": f_fis,
                 "Renta operativa $": fis["operativa"], "% Renta": pct(fis["operativa"], f_fis)},
                {"Canal": "TOTAL", "Facturación $": f_tot,
                 "Renta operativa $": op_total, "% Renta": pct(op_total, f_tot)},
            ])
            st.dataframe(
                tcons.style.format({"Facturación $": "{:,.2f}", "Renta operativa $": "{:,.2f}",
                                    "% Renta": "{:.2%}"}),
                use_container_width=True, hide_index=True)
            st.caption("Facturación = ventas brutas con IVA (antes de NC). % Renta = renta operativa / facturación. "
                       "La operativa ML ya está neta de comisión, publicidad (ADS), asesoría, mantenimiento, "
                       "devoluciones y envíos ME2/ME1/FLEX; la física, neta de cobranza de tarjeta.")

        # -------- Descarga Excel --------
        st.download_button(
            "⬇️  Descargar Excel", data=build_excel(r, fis=fis, op_total=op_total),
            file_name="Renta_ML_y_Fisico.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

        with st.expander("Detalle de envíos y controles"):
            st.write(f"**ME2/Colecta (ML):** ${r['me2']:,.2f} (columna Valor del cargo)")
            st.write(f"**DAC:** facturación ${r['dac']['total']:,.2f} − bonificación "
                     f"${r['dacnc']['total']:,.2f} = neto ${r['dac_neto']:,.2f}  ·  "
                     f"salida ventas ${r['dac']['salida']:,.2f} / proveedores ${r['dac']['entrada']:,.2f}")
            st.write(f"**FLEX Distrilogic:** ${r['flex']:,.2f} con IVA "
                     f"(neto ${flex['neto']:,.2f})")
            if fis is not None:
                st.write(f"**Canal físico:** venta ${fis['total_venta']:,.2f} · "
                         f"costo ${fis['total_costo']:,.2f} · cobranza ${fis['total_cobranza']:,.2f} · "
                         f"operativa ${fis['operativa']:,.2f}")
                if fis["nc_rev"]:
                    det = ", ".join(f"{m} doc {d}" if d else f"{m} (sin match ${v:,.0f})"
                                    for m, d, v in fis["nc_rev"])
                    st.write(f"**NC físicas neteadas:** {det}")
    except Exception as e:
        st.error(f"Error procesando los archivos: {e}")
        st.info("Verificá que cada archivo esté en el casillero correcto y con el formato esperado.")
